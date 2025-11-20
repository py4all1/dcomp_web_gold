from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum
from core.models import Empresa
from .models import NotaFiscalNacional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os
from decimal import Decimal


@login_required
def emitir_nfse(request):
    """Tela para emissão de NFS-e Nacional com abas"""
    empresa_contratante = request.user.profile.empresa
    
    # Buscar empresas emissoras
    empresas = Empresa.objects.filter(
        empresa_contratante=empresa_contratante,
        ativo=True
    )
    
    # Processar ações POST
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'importar':
            return importar_planilha(request, empresas)
        elif acao == 'consultar':
            return consultar_notas_api(request)
        elif acao == 'relatorio':
            return gerar_relatorio(request)
    
    # Filtros GET
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    empresa_filtro = request.GET.get('empresa_id')
    
    # Query de notas
    notas = NotaFiscalNacional.objects.filter(
        empresa__empresa_contratante=empresa_contratante
    )
    
    # Aplicar filtros
    if data_inicio:
        notas = notas.filter(data_emissao__gte=data_inicio)
    if data_fim:
        notas = notas.filter(data_emissao__lte=data_fim)
    if status:
        notas = notas.filter(status_nfse=status)
    if empresa_filtro:
        notas = notas.filter(empresa_id=empresa_filtro)
    
    # Se não houver filtro de data, mostrar apenas pendentes
    if not data_inicio and not data_fim:
        notas = notas.filter(status_nfse='pendente')
    
    context = {
        'user': request.user,
        'empresas': empresas,
        'notas': notas.order_by('-data_importacao'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status': status,
        'empresa_filtro': empresa_filtro,
    }
    return render(request, 'nfse_nacional/emitir.html', context)


@login_required
def gerar_modelo(request):
    """Gera arquivo Excel modelo profissional para importação de Nota Nacional de Serviço"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nota Nacional de Serviço"
    
    # Título principal
    ws.merge_cells('A1:AM1')
    titulo = ws['A1']
    titulo.value = '📋 MODELO DE IMPORTAÇÃO - NOTA NACIONAL DE SERVIÇO'
    titulo.font = Font(bold=True, size=18, color='FFFFFF', name='Calibri')
    titulo.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Subtítulo com informações
    ws.merge_cells('A2:AE2')
    subtitulo = ws['A2']
    subtitulo.value = '🏢 Sistema Tax Gold - Emissão de Notas Fiscais de Serviço Eletrônica Nacional'
    subtitulo.font = Font(italic=True, size=11, color='FFFFFF', name='Calibri')
    subtitulo.fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
    subtitulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # Cabeçalhos organizados por seção
    headers = [
        # TOMADOR (4 colunas) - Azul
        'CNPJ_CPF_TOMADOR', 'NOME_TOMADOR', 'INSCRICAO_MUNICIPAL', 'EMAIL_TOMADOR',
        # ENDEREÇO (7 colunas) - Cinza
        'LOGRADOURO', 'NUMERO', 'COMPLEMENTO', 'BAIRRO', 'CIDADE', 'UF', 'CEP',
        # SERVIÇO (8 colunas) - Laranja
        'DATA_EMISSAO', 'COD_SERVICO', 'COD_TRIBUTACAO_MUNICIPIO', 'DESCRICAO',
        'VALOR_TOTAL', 'DEDUCOES', 'DESCONTO_INCONDICIONADO', 'DESCONTO_CONDICIONADO',
        # ISS (4 colunas) - Roxo
        'ALIQUOTA_ISS', 'TIPO_TRIBUTACAO', 'ISS_RETIDO', 'MUNICIPIO_INCIDENCIA',
        # RETENÇÕES (5 colunas) - Vermelho
        'PIS_RETIDO', 'COFINS_RETIDO', 'IRRF_RETIDO', 'CSLL_RETIDO', 'INSS_RETIDO',
        # RPS E OBS (3 colunas) - Cinza Escuro
        'NUMERO_RPS', 'SERIE_RPS', 'OBSERVACOES'
    ]
    
    # Cores por seção
    cores_secoes = {
        'tomador': '4472C4',      # Azul
        'endereco': '7F7F7F',     # Cinza
        'servico': 'ED7D31',      # Laranja
        'iss': '9966CC',          # Roxo
        'retencoes': 'C00000',    # Vermelho
        'rps': '595959'           # Cinza Escuro
    }
    
    # Aplicar cores por seção
    secoes = [
        (1, 4, 'tomador'),
        (5, 11, 'endereco'),
        (12, 19, 'servico'),
        (20, 23, 'iss'),
        (24, 28, 'retencoes'),
        (29, 31, 'rps')
    ]
    
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Aplicar estilo ao cabeçalho com cores por seção
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Linha 4 para cabeçalhos
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        
        # Aplicar cor da seção
        for inicio, fim, secao in secoes:
            if inicio <= col_num <= fim:
                cell.fill = PatternFill(start_color=cores_secoes[secao], 
                                       end_color=cores_secoes[secao], 
                                       fill_type='solid')
                break
    
    # Ajustar altura da linha do cabeçalho
    ws.row_dimensions[4].height = 40
    
    # Adicionar legenda de seções (linha 3)
    ws.merge_cells('A3:D3')
    ws['A3'].value = '👤 TOMADOR'
    ws['A3'].fill = PatternFill(start_color='D6E4F5', end_color='D6E4F5', fill_type='solid')
    ws['A3'].font = Font(bold=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('E3:K3')
    ws['E3'].value = '📍 ENDEREÇO'
    ws['E3'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    ws['E3'].font = Font(bold=True, size=9)
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('L3:S3')
    ws['L3'].value = '📋 SERVIÇO'
    ws['L3'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    ws['L3'].font = Font(bold=True, size=9)
    ws['L3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('T3:W3')
    ws['T3'].value = '🏛️ ISS'
    ws['T3'].fill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
    ws['T3'].font = Font(bold=True, size=9)
    ws['T3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('X3:AB3')
    ws['X3'].value = '💰 RETENÇÕES'
    ws['X3'].fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    ws['X3'].font = Font(bold=True, size=9)
    ws['X3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('AC3:AE3')
    ws['AC3'].value = '📝 RPS/OBS'
    ws['AC3'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['AC3'].font = Font(bold=True, size=9)
    ws['AC3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[3].height = 25
    
    # Linha de exemplo (linha 5)
    exemplo = [
        '12345678000190', 'Empresa Tomadora Ltda', '123456', 'contato@empresa.com.br',
        'Av. Paulista', '1000', 'Sala 100', 'Bela Vista', 'São Paulo', 'SP', '01310-100',
        '2024-11-20', '01.07', '010701', 'Serviços de consultoria empresarial e gestão estratégica',
        '10000.00', '0.00', '0.00', '0.00',
        '5.00', 'T', 'NAO', 'São Paulo',
        '0.00', '0.00', '0.00', '0.00', '0.00',
        '001', '1', 'Serviço prestado conforme contrato'
    ]
    
    # Adicionar linha de exemplo com formatação alternada
    for col_num, valor in enumerate(exemplo, 1):
        cell = ws.cell(row=5, column=col_num, value=valor)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(name='Calibri', size=10, italic=True, color='0066CC')
        # Fundo alternado suave
        cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    
    ws.row_dimensions[5].height = 25
    
    # Ajustar largura das colunas
    column_widths = [
        18, 30, 18, 30,  # Tomador
        25, 10, 15, 15, 20, 5, 12,  # Endereço
        15, 12, 20, 50,  # Serviço
        15, 12, 20, 20,  # Valores
        12, 18, 12, 20,  # ISS
        12, 12, 12, 12, 12,  # Retenções
        15, 10, 40  # RPS e Obs
    ]
    
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Bordas para cabeçalhos e exemplo
    thin_border = Border(
        left=Side(style='thin', color='999999'),
        right=Side(style='thin', color='999999'),
        top=Side(style='thin', color='999999'),
        bottom=Side(style='thin', color='999999')
    )
    
    medium_border = Border(
        left=Side(style='medium', color='333333'),
        right=Side(style='medium', color='333333'),
        top=Side(style='medium', color='333333'),
        bottom=Side(style='medium', color='333333')
    )
    
    # Aplicar bordas nos cabeçalhos (linha 4)
    for col in range(1, len(headers) + 1):
        ws.cell(row=4, column=col).border = medium_border
    
    # Aplicar bordas na linha de exemplo (linha 5)
    for col in range(1, len(headers) + 1):
        ws.cell(row=5, column=col).border = thin_border
    
    # Borda no título
    ws['A1'].border = Border(
        bottom=Side(style='thick', color='1F4E78')
    )
    
    # Congelar painéis (fixar cabeçalhos)
    ws.freeze_panes = 'A5'
    
    # ========== CRIAR ABA DE INSTRUÇÕES ==========
    ws_instrucoes = wb.create_sheet("📋 Instruções")
    
    # Título da aba de instruções
    ws_instrucoes.merge_cells('A1:F1')
    titulo_inst = ws_instrucoes['A1']
    titulo_inst.value = '📋 INSTRUÇÕES DE PREENCHIMENTO'
    titulo_inst.font = Font(bold=True, size=16, color='FFFFFF', name='Calibri')
    titulo_inst.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    titulo_inst.alignment = Alignment(horizontal='center', vertical='center')
    ws_instrucoes.row_dimensions[1].height = 30
    
    # Conteúdo das instruções
    instrucoes_texto = [
        ("", ""),
        ("📌 CAMPOS OBRIGATÓRIOS", ""),
        ("CNPJ_CPF_TOMADOR", "CNPJ ou CPF do tomador do serviço"),
        ("NOME_TOMADOR", "Razão social ou nome completo"),
        ("DATA_EMISSAO", "Data de emissão no formato AAAA-MM-DD (ex: 2024-11-20)"),
        ("COD_SERVICO", "Código do serviço conforme tabela municipal"),
        ("DESCRICAO", "Descrição detalhada do serviço prestado"),
        ("VALOR_TOTAL", "Valor total do serviço em R$ (ex: 1000.00)"),
        ("ALIQUOTA_ISS", "Alíquota do ISS em % (ex: 5.00 para 5%)"),
        ("", ""),
        ("📝 TIPO DE TRIBUTAÇÃO", ""),
        ("T", "Tributado no município"),
        ("F", "Fora do município (serviço prestado em outro município)"),
        ("I", "Isento de ISS"),
        ("N", "Não tributado"),
        ("", ""),
        ("💰 RETENÇÕES", ""),
        ("ISS_RETIDO", "SIM ou NAO - Indica se o ISS foi retido pelo tomador"),
        ("PIS_RETIDO", "Valor do PIS retido (ex: 65.00)"),
        ("COFINS_RETIDO", "Valor do COFINS retido (ex: 300.00)"),
        ("IRRF_RETIDO", "Valor do IRRF retido (ex: 150.00)"),
        ("CSLL_RETIDO", "Valor do CSLL retido (ex: 100.00)"),
        ("INSS_RETIDO", "Valor do INSS retido (ex: 110.00)"),
        ("", ""),
        ("📋 OBSERVAÇÕES IMPORTANTES", ""),
        ("Valores", "Sempre use ponto como separador decimal (ex: 1000.00)"),
        ("Alíquotas", "Sempre em percentual com 2 decimais (ex: 5.00 = 5%)"),
        ("Datas", "Formato obrigatório: AAAA-MM-DD (ex: 2024-11-20)"),
        ("SIM/NAO", "Use sempre maiúsculas: SIM ou NAO"),
        ("CEP", "Apenas números, sem hífen (ex: 01310100)"),
        ("", ""),
        ("⚠️ REFORMA TRIBUTÁRIA", ""),
        ("IBS e CBS", "Os campos de IBS e CBS foram removidos pois a Reforma Tributária"),
        ("", "só entra em vigor oficialmente a partir de 2026."),
        ("", "Para serviços, não há incidência de IBS, apenas CBS."),
    ]
    
    linha_atual = 3
    for campo, descricao in instrucoes_texto:
        if campo == "" and descricao == "":
            linha_atual += 1
            continue
        
        cell_campo = ws_instrucoes.cell(row=linha_atual, column=1, value=campo)
        cell_desc = ws_instrucoes.cell(row=linha_atual, column=2, value=descricao)
        
        # Formatação para títulos de seção
        if descricao == "" and campo != "":
            ws_instrucoes.merge_cells(f'A{linha_atual}:F{linha_atual}')
            cell_campo.font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
            cell_campo.fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
            cell_campo.alignment = Alignment(horizontal='left', vertical='center')
            ws_instrucoes.row_dimensions[linha_atual].height = 25
        else:
            cell_campo.font = Font(bold=True, size=10, name='Calibri')
            cell_desc.font = Font(size=10, name='Calibri')
            cell_campo.alignment = Alignment(horizontal='left', vertical='center')
            cell_desc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws_instrucoes.row_dimensions[linha_atual].height = 20
        
        linha_atual += 1
    
    # Ajustar larguras das colunas da aba de instruções
    ws_instrucoes.column_dimensions['A'].width = 25
    ws_instrucoes.column_dimensions['B'].width = 60
    
    # Preparar resposta com nome profissional
    from datetime import datetime
    data_atual = datetime.now().strftime('%Y%m%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TaxGold_Modelo_NotaNacionalServico_{data_atual}.xlsx"'
    wb.save(response)
    
    return response


def importar_planilha(request, empresas):
    """Importa planilha Excel com notas fiscais nacionais"""
    try:
        empresa_id = request.POST.get('empresa_id')
        arquivo = request.FILES.get('arquivo_excel')
        
        if not empresa_id or not arquivo:
            messages.error(request, 'Selecione a empresa e o arquivo Excel.')
            return redirect('nfse_nacional:emitir')
        
        empresa = get_object_or_404(Empresa, id=empresa_id)
        
        # Validar extensão
        if not arquivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Arquivo deve ser .xlsx ou .xls')
            return redirect('nfse_nacional:emitir')
        
        # Ler planilha
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        importadas = 0
        erros = []
        
        # Pular cabeçalho (linha 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Validar se linha não está vazia
                if not any(row):
                    continue
                
                # Converter data
                data_emissao = row[11]
                if isinstance(data_emissao, str):
                    data_emissao = datetime.strptime(data_emissao, '%Y-%m-%d').date()
                elif isinstance(data_emissao, datetime):
                    data_emissao = data_emissao.date()
                
                # Criar nota
                nota = NotaFiscalNacional(
                    empresa=empresa,
                    cnpj_contribuinte=empresa.cnpj,  # CNPJ da empresa selecionada
                    # Tomador
                    cnpj_cpf_tomador=str(row[0] or '').strip(),
                    nome_tomador=str(row[1] or '').strip(),
                    inscricao_municipal_tomador=str(row[2] or '').strip() if row[2] else None,
                    email_tomador=str(row[3] or '').strip() if row[3] else None,
                    # Endereço
                    logradouro_tomador=str(row[4] or '').strip() if row[4] else None,
                    numero_tomador=str(row[5] or '').strip() if row[5] else None,
                    complemento_tomador=str(row[6] or '').strip() if row[6] else None,
                    bairro_tomador=str(row[7] or '').strip() if row[7] else None,
                    cidade_tomador=str(row[8] or '').strip() if row[8] else None,
                    uf_tomador=str(row[9] or '').strip() if row[9] else None,
                    cep_tomador=str(row[10] or '').strip() if row[10] else None,
                    # Serviço
                    data_emissao=data_emissao,
                    cod_servico=str(row[12] or '').strip(),
                    cod_tributacao_municipio=str(row[13] or '').strip() if row[13] else None,
                    descricao=str(row[14] or '').strip(),
                    valor_total=Decimal(str(row[15] or 0)),
                    deducoes=Decimal(str(row[16] or 0)),
                    desconto_incondicionado=Decimal(str(row[17] or 0)),
                    desconto_condicionado=Decimal(str(row[18] or 0)),
                    # ISS
                    aliquota_iss=Decimal(str(row[19] or 0)),
                    tipo_tributacao=str(row[20] or 'T').strip()[0].upper(),
                    iss_retido=str(row[21] or '').strip().upper() in ['SIM', 'S', 'TRUE', '1'],
                    municipio_incidencia=str(row[22] or '').strip() if row[22] else None,
                    # Retenções (índices ajustados após remoção de IBS/CBS)
                    pis_retido=Decimal(str(row[23] or 0)),
                    cofins_retido=Decimal(str(row[24] or 0)),
                    irrf_retido=Decimal(str(row[25] or 0)),
                    csll_retido=Decimal(str(row[26] or 0)),
                    inss_retido=Decimal(str(row[27] or 0)),
                    # RPS
                    numero_rps=str(row[28] or '').strip() if row[28] else None,
                    serie_rps=str(row[29] or '').strip() if row[29] else None,
                    observacoes=str(row[30] or '').strip() if row[30] else None,
                    status_nfse='pendente'
                )
                nota.save()
                importadas += 1
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        if importadas > 0:
            messages.success(request, f'{importadas} nota(s) importada(s) com sucesso!')
        
        if erros:
            messages.warning(request, f'{len(erros)} erro(s) encontrado(s). Verifique os dados.')
        
    except Exception as e:
        messages.error(request, f'Erro ao importar planilha: {str(e)}')
    
    return redirect('nfse_nacional:emitir')


@login_required
def emitir_notas(request):
    """Emite notas selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma nota selecionada.'
            })
        
        # Processar emissão
        from .services import ProcessadorNFSeNacional
        resultados = []
        emitidas = 0
        erros = 0
        
        for nota_id in notas_ids:
            try:
                nota = NotaFiscalNacional.objects.get(id=nota_id)
                
                # Verificar se já foi emitida
                if nota.status_nfse in ['emitida', 'cancelada']:
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_documento': nota.numero_rps or str(nota.id),
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ Nota {nota.numero_rps or nota.id}\n'\
                                   f'Não pode ser emitida:\n\n'\
                                   f'Status atual: {nota.get_status_nfse_display()}\n'\
                                   f'Tomador: {nota.nome_tomador}\n'\
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    erros += 1
                    continue
                
                # Emitir nota
                processador = ProcessadorNFSeNacional(nota.empresa)
                resultado = processador.emitir_nota(nota)
                
                if resultado['sucesso']:
                    emitidas += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': True,
                        'numero_documento': nota.numero_rps or str(nota.id),
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'✅ Nota {nota.numero_rps or nota.id}\n'\
                                   f'Emitida com sucesso!\n'\
                                   f'Tomador: {nota.nome_tomador}\n'\
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                else:
                    erros += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_documento': nota.numero_rps or str(nota.id),
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': resultado.get('mensagem', 'Erro desconhecido')
                    })
                    
            except Exception as e:
                erros += 1
                resultados.append({
                    'id': nota_id,
                    'sucesso': False,
                    'mensagem': f'Erro: {str(e)}'
                })
        
        # Mensagem de resumo profissional
        if emitidas > 0 and erros == 0:
            mensagem_resumo = f'🎉 Emissão concluída com sucesso!\n\nTodas as {emitidas} nota(s) foram emitidas corretamente.'
        elif emitidas > 0 and erros > 0:
            mensagem_resumo = f'⚠️ Emissão concluída com avisos.\n\n{emitidas} nota(s) emitida(s) com sucesso.\n{erros} nota(s) com erro.'
        else:
            mensagem_resumo = f'❌ Não foi possível emitir as notas.\n\n{erros} erro(s) encontrado(s).'
        
        return JsonResponse({
            'sucesso': emitidas > 0,
            'mensagem': mensagem_resumo,
            'emitidas': emitidas,
            'erros': erros,
            'total': len(notas_ids),
            'resultados': resultados
        })
    
    return JsonResponse({'sucesso': False, 'mensagem': 'Método não permitido'})


@login_required
def excluir_notas(request):
    """Exclui notas não emitidas selecionadas (pendentes, com erro ou sem status)"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma nota selecionada.'
            })
        
        # Buscar informações das notas antes de excluir
        from django.db.models import Q
        notas_para_excluir = NotaFiscalNacional.objects.filter(
            id__in=notas_ids
        ).filter(
            Q(status_nfse='pendente') | 
            Q(status_nfse='erro') | 
            Q(status_nfse__isnull=True) | 
            Q(status_nfse='')
        )
        
        resultados = []
        for nota in notas_para_excluir:
            resultados.append({
                'id': nota.id,
                'numero_documento': nota.numero_rps or str(nota.id),
                'tomador': nota.nome_tomador,
                'valor': float(nota.valor_total),
                'status': nota.get_status_nfse_display() if nota.status_nfse else 'Sem Status'
            })
        
        # Excluir notas
        excluidas = notas_para_excluir.delete()[0]
        
        if excluidas > 0:
            mensagem_resumo = f'🎉 Exclusão concluída com sucesso!\n\n{excluidas} nota(s) excluída(s).'
            return JsonResponse({
                'sucesso': True,
                'mensagem': mensagem_resumo,
                'excluidas': excluidas,
                'total': len(notas_ids),
                'resultados': resultados
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'mensagem': '❌ Nenhuma nota foi excluída.\n\nApenas notas pendentes, com erro ou sem status podem ser excluídas.'
            })
    
    return JsonResponse({'sucesso': False, 'mensagem': 'Método não permitido'})


@login_required
def editar_nota(request, nota_id):
    """Edita uma nota fiscal pendente"""
    nota = get_object_or_404(NotaFiscalNacional, id=nota_id)
    
    # Verificar se a nota pode ser editada (apenas pendentes ou com erro)
    if nota.status_nfse not in ['pendente', 'erro']:
        messages.error(request, 'Apenas notas pendentes ou com erro podem ser editadas.')
        return redirect('nfse_nacional:emitir')
    
    if request.method == 'POST':
        try:
            # Atualizar dados do tomador
            nota.cnpj_cpf_tomador = request.POST.get('cnpj_cpf_tomador', '').strip()
            nota.nome_tomador = request.POST.get('nome_tomador', '').strip()
            nota.inscricao_municipal_tomador = request.POST.get('inscricao_municipal_tomador', '').strip() or None
            nota.email_tomador = request.POST.get('email_tomador', '').strip() or None
            
            # Endereço
            nota.logradouro_tomador = request.POST.get('logradouro_tomador', '').strip() or None
            nota.numero_tomador = request.POST.get('numero_tomador', '').strip() or None
            nota.complemento_tomador = request.POST.get('complemento_tomador', '').strip() or None
            nota.bairro_tomador = request.POST.get('bairro_tomador', '').strip() or None
            nota.cidade_tomador = request.POST.get('cidade_tomador', '').strip() or None
            nota.uf_tomador = request.POST.get('uf_tomador', '').strip() or None
            nota.cep_tomador = request.POST.get('cep_tomador', '').strip() or None
            
            # Serviço
            data_emissao_str = request.POST.get('data_emissao')
            if data_emissao_str:
                nota.data_emissao = datetime.strptime(data_emissao_str, '%Y-%m-%d').date()
            
            nota.cod_servico = request.POST.get('cod_servico', '').strip()
            nota.cod_tributacao_municipio = request.POST.get('cod_tributacao_municipio', '').strip() or None
            nota.descricao = request.POST.get('descricao', '').strip()
            nota.valor_total = Decimal(request.POST.get('valor_total', '0'))
            nota.deducoes = Decimal(request.POST.get('deducoes', '0'))
            nota.desconto_incondicionado = Decimal(request.POST.get('desconto_incondicionado', '0'))
            nota.desconto_condicionado = Decimal(request.POST.get('desconto_condicionado', '0'))
            
            # ISS
            nota.aliquota_iss = Decimal(request.POST.get('aliquota_iss', '0'))
            nota.tipo_tributacao = request.POST.get('tipo_tributacao', 'T').strip()[0].upper()
            nota.iss_retido = request.POST.get('iss_retido') == 'on'
            nota.municipio_incidencia = request.POST.get('municipio_incidencia', '').strip() or None
            
            # Retenções
            nota.pis_retido = Decimal(request.POST.get('pis_retido', '0'))
            nota.cofins_retido = Decimal(request.POST.get('cofins_retido', '0'))
            nota.irrf_retido = Decimal(request.POST.get('irrf_retido', '0'))
            nota.csll_retido = Decimal(request.POST.get('csll_retido', '0'))
            nota.inss_retido = Decimal(request.POST.get('inss_retido', '0'))
            
            # RPS
            nota.numero_rps = request.POST.get('numero_rps', '').strip() or None
            nota.serie_rps = request.POST.get('serie_rps', '').strip() or None
            nota.observacoes = request.POST.get('observacoes', '').strip() or None
            
            # Limpar mensagem de erro se estava com erro
            if nota.status_nfse == 'erro':
                nota.mensagem_erro = None
                nota.status_nfse = 'pendente'
            
            nota.save()
            messages.success(request, 'Nota atualizada com sucesso!')
            return redirect('nfse_nacional:emitir')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar nota: {str(e)}')
    
    context = {
        'nota': nota,
        'user': request.user,
    }
    return render(request, 'nfse_nacional/editar_nota.html', context)


@login_required
def cancelar_notas(request):
    """Cancela notas emitidas selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            messages.error(request, 'Nenhuma nota selecionada.')
            return redirect('nfse_nacional:emitir')
        
        # TODO: Implementar cancelamento via API
        messages.info(request, f'Cancelamento de {len(notas_ids)} nota(s) em desenvolvimento.')
    
    return redirect('nfse_nacional:emitir')


@login_required
def salvar_pdfs(request):
    """Salva PDFs das notas selecionadas"""
    notas_ids = request.GET.get('notas', '').split(',')
    
    if not notas_ids or notas_ids == ['']:
        messages.error(request, 'Nenhuma nota selecionada.')
        return redirect('nfse_nacional:emitir')
    
    # TODO: Implementar download de PDFs
    messages.info(request, f'Download de PDFs em desenvolvimento.')
    return redirect('nfse_nacional:emitir')


@login_required
def detalhes_nota(request, nota_id):
    """Retorna HTML com detalhes da nota"""
    nota = get_object_or_404(NotaFiscalNacional, id=nota_id)
    
    html = f"""
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Tomador</h6>
            <p><strong>CNPJ/CPF:</strong> {nota.cnpj_cpf_tomador}</p>
            <p><strong>Nome:</strong> {nota.nome_tomador}</p>
            <p><strong>Insc. Municipal:</strong> {nota.inscricao_municipal_tomador or '-'}</p>
            <p><strong>E-mail:</strong> {nota.email_tomador or '-'}</p>
            <p><strong>Endereço:</strong> {nota.logradouro_tomador or ''} {nota.numero_tomador or ''}</p>
            <p><strong>Cidade/UF:</strong> {nota.cidade_tomador or '-'}/{nota.uf_tomador or '-'}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Serviço</h6>
            <p><strong>Data Emissão:</strong> {nota.data_emissao.strftime('%d/%m/%Y')}</p>
            <p><strong>Código:</strong> {nota.cod_servico}</p>
            <p><strong>Descrição:</strong> {nota.descricao}</p>
            <p><strong>Valor Total:</strong> R$ {nota.valor_total}</p>
            <p><strong>Deduções:</strong> R$ {nota.deducoes}</p>
        </div>
    </div>
    <hr>
    <div class="row">
        <div class="col-md-4">
            <h6 class="text-primary">Tributação Municipal (ISS)</h6>
            <p><strong>Alíquota ISS:</strong> {nota.aliquota_iss}%</p>
            <p><strong>Valor ISS:</strong> R$ {nota.valor_iss:.2f}</p>
            <p><strong>ISS Retido:</strong> {'Sim' if nota.iss_retido else 'Não'}</p>
            <p><strong>Tipo:</strong> {nota.get_tipo_tributacao_display()}</p>
        </div>
        <div class="col-md-4">
            <h6 class="text-success">Reforma Tributária - IBS</h6>
            <p><strong>IBS Devido:</strong> {'Sim' if nota.ibs_devido else 'Não'}</p>
            <p><strong>Alíquota IBS:</strong> {nota.aliquota_ibs}%</p>
            <p><strong>Valor IBS:</strong> R$ {nota.valor_ibs}</p>
            <p><strong>IBS Retido:</strong> {'Sim' if nota.ibs_retido else 'Não'}</p>
        </div>
        <div class="col-md-4">
            <h6 class="text-info">Reforma Tributária - CBS</h6>
            <p><strong>CBS Devido:</strong> {'Sim' if nota.cbs_devido else 'Não'}</p>
            <p><strong>Alíquota CBS:</strong> {nota.aliquota_cbs}%</p>
            <p><strong>Valor CBS:</strong> R$ {nota.valor_cbs}</p>
            <p><strong>CBS Retido:</strong> {'Sim' if nota.cbs_retido else 'Não'}</p>
        </div>
    </div>
    <hr>
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Retenções Federais</h6>
            <p><strong>PIS:</strong> R$ {nota.pis_retido}</p>
            <p><strong>COFINS:</strong> R$ {nota.cofins_retido}</p>
            <p><strong>IRRF:</strong> R$ {nota.irrf_retido}</p>
            <p><strong>CSLL:</strong> R$ {nota.csll_retido}</p>
            <p><strong>INSS:</strong> R$ {nota.inss_retido}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Controle</h6>
            <p><strong>Status:</strong> {nota.get_status_nfse_display()}</p>
            <p><strong>RPS:</strong> {nota.numero_rps or '-'} / Série: {nota.serie_rps or '-'}</p>
            <p><strong>NFS-e:</strong> {nota.numero_nfse or '-'}</p>
            <p><strong>Valor Líquido:</strong> R$ {nota.valor_liquido:.2f}</p>
        </div>
    </div>
    """
    
    return HttpResponse(html)


def consultar_notas_api(request):
    """Consulta notas via API"""
    # TODO: Implementar integração com API
    messages.info(request, 'Consulta de notas via API em desenvolvimento.')
    return redirect('nfse_nacional:emitir')


def gerar_relatorio(request):
    """Gera relatório conforme tipo selecionado"""
    # TODO: Implementar geração de relatórios
    messages.info(request, 'Geração de relatórios em desenvolvimento.')
    return redirect('nfse_nacional:emitir')
