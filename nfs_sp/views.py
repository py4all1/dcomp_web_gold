from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum
from core.models import Empresa, NotaFiscalSP, NotaFiscalTomadorSP
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import os
from decimal import Decimal
import json

# Importar serviços de NFS-e
from nfs_sp.services.nfe_eventos_django import EventoNFeDjango
from nfs_sp.services.processador_django import ProcessadorNFeDjango


@login_required
def emitir_nfs(request):
    """Tela para emissão de NFS-e São Paulo com abas"""
    # Pegar empresa contratante do usuário
    empresa_contratante = request.user.profile.empresa if hasattr(request.user, 'profile') else None
    
    if not empresa_contratante:
        messages.error(request, 'Usuário sem empresa vinculada.')
        return redirect('core:home')
    
    # Listar empresas cadastradas
    empresas = Empresa.objects.filter(
        empresa_contratante=empresa_contratante,
        ativo=True
    )
    
    # Processar ações POST
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'importar':
            return importar_planilha(request, empresas)
        elif acao == 'importar_nfts':
            return importar_planilha_nfts(request, empresas)
        elif acao == 'consultar':
            return consultar_notas_api(request)
        elif acao == 'relatorio':
            return gerar_relatorio(request)
    
    # Filtros GET
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    cnpj_contribuinte = request.GET.get('cnpj_contribuinte')
    aba_ativa = request.GET.get('aba', 'emitir')  # Identifica qual aba está ativa
    
    # Query de notas NFS-e
    notas = NotaFiscalSP.objects.filter(
        empresa__empresa_contratante=empresa_contratante
    )
    
    # Aplicar filtros APENAS se a aba ativa for "emitir"
    if aba_ativa == 'emitir':
        # Filtro por CNPJ do contribuinte (empresa específica)
        if cnpj_contribuinte:
            # Filtra por cnpj_contribuinte OU por empresa.cnpj (para notas antigas)
            notas = notas.filter(
                Q(cnpj_contribuinte=cnpj_contribuinte) | 
                Q(empresa__cnpj=cnpj_contribuinte)
            )
        
        if data_inicio:
            # Filtra por data_emissao OU data_importacao (para notas sem data_emissao)
            notas = notas.filter(
                Q(data_emissao__gte=data_inicio) | 
                Q(data_emissao__isnull=True, data_importacao__date__gte=data_inicio)
            )
        if data_fim:
            notas = notas.filter(
                Q(data_emissao__lte=data_fim) | 
                Q(data_emissao__isnull=True, data_importacao__date__lte=data_fim)
            )
        if status:
            notas = notas.filter(status_rps=status)
        
        # Se não houver nenhum filtro, mostrar apenas pendentes
        if not data_inicio and not data_fim and not status and not cnpj_contribuinte:
            notas = notas.filter(status_rps='pendente')
    else:
        # Se a aba não for "emitir", mostrar apenas pendentes por padrão
        notas = notas.filter(status_rps='pendente')
    
    # Query de NFTS
    notas_nfts = NotaFiscalTomadorSP.objects.filter(
        empresa__empresa_contratante=empresa_contratante
    )
    
    # Aplicar filtros APENAS se a aba ativa for "emitir-nfts"
    if aba_ativa == 'emitir-nfts':
        # Filtro por CNPJ do contribuinte (empresa específica)
        if cnpj_contribuinte:
            # Filtra por cnpj_contribuinte OU por empresa.cnpj (para notas antigas)
            notas_nfts = notas_nfts.filter(
                Q(cnpj_contribuinte=cnpj_contribuinte) | 
                Q(empresa__cnpj=cnpj_contribuinte)
            )
        
        if data_inicio:
            # Filtra por data_prestacao_servico OU data_importacao (para notas sem data)
            notas_nfts = notas_nfts.filter(
                Q(data_prestacao_servico__gte=data_inicio) | 
                Q(data_prestacao_servico__isnull=True, data_importacao__date__gte=data_inicio)
            )
        if data_fim:
            notas_nfts = notas_nfts.filter(
                Q(data_prestacao_servico__lte=data_fim) | 
                Q(data_prestacao_servico__isnull=True, data_importacao__date__lte=data_fim)
            )
        if status:
            notas_nfts = notas_nfts.filter(status_nfts=status)
        
        # Se não houver nenhum filtro, mostrar apenas pendentes
        if not data_inicio and not data_fim and not status and not cnpj_contribuinte:
            notas_nfts = notas_nfts.filter(status_nfts='pendente')
    else:
        # Se a aba não for "emitir-nfts", mostrar apenas pendentes por padrão
        notas_nfts = notas_nfts.filter(status_nfts='pendente')
    
    context = {
        'user': request.user,
        'empresas': empresas,
        'notas': notas.order_by('-data_importacao'),
        'notas_nfts': notas_nfts.order_by('-data_importacao'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status': status,
        'cnpj_contribuinte': cnpj_contribuinte,
    }
    return render(request, 'nfs_sp/emitir.html', context)


def importar_planilha(request, empresas):
    """Importa planilha Excel com notas fiscais"""
    try:
        empresa_id = request.POST.get('empresa_id')
        arquivo = request.FILES.get('arquivo_excel')
        
        if not empresa_id or not arquivo:
            messages.error(request, 'Selecione a empresa e o arquivo Excel.')
            return redirect('nfs_sp:emitir')
        
        empresa = get_object_or_404(Empresa, id=empresa_id)
        
        # Validar extensão
        if not arquivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Arquivo deve ser .xlsx ou .xls')
            return redirect('nfs_sp:emitir')
        
        # Ler planilha
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        importadas = 0
        erros = []
        
        # Pular cabeçalho (linha 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Validar se linha não está vazia
                if not any(row):
                    continue
                
                # Extrair dados (conforme colunas especificadas)
                nota = NotaFiscalSP(
                    empresa=empresa,
                    cnpj_contribuinte=empresa.cnpj,  # CNPJ da empresa prestadora
                    cnpj_cpf_tomador=str(row[1] or '').strip(),
                    nome_tomador=str(row[2] or '').strip(),
                    cep_tomador=str(row[3] or '').strip() if row[3] else None,
                    logradouro_tomador=str(row[4] or '').strip() if row[4] else None,
                    numero_tomador=str(row[5] or '').strip() if row[5] else None,
                    bairro_tomador=str(row[6] or '').strip() if row[6] else None,
                    cidade_tomador=str(row[7] or '').strip() if row[7] else None,
                    uf_tomador=str(row[8] or '').strip() if row[8] else None,
                    email_tomador=str(row[9] or '').strip() if row[9] else None,
                    cod_servico=str(row[10] or '').strip(),
                    descricao=str(row[11] or '').strip(),
                    valor_total=Decimal(str(row[12] or 0)),
                    deducoes=Decimal(str(row[13] or 0)),
                    aliquota=Decimal(str(row[14] or 0)),
                    tipo_tributacao=str(row[15] or 'T').strip()[0].upper(),
                    iss_retido=str(row[17] or '').strip().upper() in ['SIM', 'S', 'TRUE', '1'],
                    pis_retido=Decimal(str(row[18] or 0)),
                    cofins_retido=Decimal(str(row[19] or 0)),
                    irrf_retido=Decimal(str(row[20] or 0)),
                    csll_retido=Decimal(str(row[21] or 0)),
                    inss_retido=Decimal(str(row[22] or 0)),
                    status_rps='pendente'
                )
                nota.save()
                importadas += 1
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        if importadas > 0:
            messages.success(request, f'{importadas} nota(s) importada(s) com sucesso!')
        
        if erros:
            messages.warning(request, f'{len(erros)} erro(s) encontrado(s). Verifique os dados.')
        
    except Exception as e:
        messages.error(request, f'Erro ao importar planilha: {str(e)}')
    
    return redirect('nfs_sp:emitir')


@login_required
def gerar_modelo(request):
    """Gera arquivo Excel modelo para importação"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo NFS-e SP"
    
    # Cabeçalhos
    headers = [
        'CNPJ_PRESTADOR', 'CNPJ_CPF_TOMADOR', 'NOME_TOMADOR', 'CEP_TOMADOR',
        'LOGRADOURO', 'NUMERO', 'BAIRRO', 'CIDADE', 'UF', 'EMAIL',
        'COD_SERVICO', 'DESCRICAO', 'VALOR_TOTAL', 'DEDUCOES', 'ALIQUOTA',
        'TIPO_TRIBUTACAO', 'STATUS_RPS', 'ISS_RETIDO', 'PIS_RETIDO',
        'COFINS_RETIDO', 'IRRF_RETIDO', 'CSLL_RETIDO', 'INSS_RETIDO'
    ]
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Linha de exemplo
    exemplo = [
        '12345678000190', '98765432000100', 'Cliente Exemplo Ltda', '01310-100',
        'Av Paulista', '1000', 'Bela Vista', 'São Paulo', 'SP', 'cliente@exemplo.com',
        '07498', 'Serviços de consultoria empresarial', '1000.00', '0.00', '2.00',
        'T', 'Pendente', 'NÃO', '0.00', '0.00', '0.00', '0.00', '0.00'
    ]
    
    for col_num, valor in enumerate(exemplo, 1):
        ws.cell(row=2, column=col_num, value=valor)
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_nfse_sp.xlsx"'
    wb.save(response)
    
    return response


@login_required
def emitir_notas(request):
    """Emite notas fiscais (RPS) selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma nota selecionada.'
            })
        
        # Processar emissão
        resultados = []
        emitidas = 0
        erros = 0
        
        for nota_id in notas_ids:
            try:
                nota = NotaFiscalSP.objects.get(id=nota_id)
                
                # Verificar se já foi emitida ou cancelada (não permite reenvio)
                if nota.status_rps in ['emitida', 'cancelada']:
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'mensagem': f'RPS {nota.numero_rps} já foi {nota.get_status_rps_display().lower()}'
                    })
                    erros += 1
                    continue
                
                # Se estava com erro, limpar mensagem de erro anterior
                if nota.status_rps == 'erro':
                    nota.mensagem_erro = None
                
                # Gerar número RPS se não existir
                if not nota.numero_rps:
                    # Buscar o último número RPS da empresa
                    ultimo_rps = NotaFiscalSP.objects.filter(
                        empresa=nota.empresa,
                        numero_rps__isnull=False
                    ).order_by('-numero_rps').first()
                    
                    if ultimo_rps and ultimo_rps.numero_rps:
                        try:
                            proximo_numero = int(ultimo_rps.numero_rps) + 1
                        except:
                            proximo_numero = 1
                    else:
                        proximo_numero = 1
                    
                    nota.numero_rps = str(proximo_numero).zfill(12)
                    nota.serie_rps = nota.serie_rps or 'RPS'
                    nota.tributacao_rps = nota.tributacao_rps or nota.tipo_tributacao
                    nota.save()
                
                # Criar XML e enviar
                evento = EventoNFeDjango(nota.empresa)
                xml = evento.criar_pedido_envio_rps(nota)
                
                processador = ProcessadorNFeDjango(nota.empresa)
                resultado = processador.enviar_rps(xml)
                
                if resultado['sucesso']:
                    # Atualizar nota
                    nota.numero_nfse = resultado.get('numero_nfe')
                    nota.codigo_verificacao = resultado.get('codigo_verificacao')
                    nota.status_rps = 'emitida'
                    nota.data_emissao = datetime.now()
                    
                    # Log para debug
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.info(f"Salvando nota {nota.id}: numero_nfse={nota.numero_nfse}, codigo={nota.codigo_verificacao}")
                    
                    # Se não tem número, logar o XML de resposta para debug
                    if not nota.numero_nfse and resultado.get('xml_resposta'):
                        logger.warning(f"Nota {nota.id} emitida mas sem numero_nfse!")
                        logger.debug(f"XML Resposta: {resultado.get('xml_resposta')[:500]}...")
                    
                    nota.save()
                    
                    emitidas += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': True,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'✅ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Emitido com sucesso!\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}',
                        'numero_nfe': nota.numero_nfse,
                        'codigo_verificacao': nota.codigo_verificacao,
                        'url': resultado.get('url_nfe')
                    })
                else:
                    nota.status_rps = 'erro'
                    nota.mensagem_erro = resultado.get('mensagem', 'Erro desconhecido')
                    nota.save()
                    
                    erros += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Erro na emissão:\n\n'
                                   f'{resultado.get("mensagem", "Erro ao emitir RPS")}\n\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    
            except Exception as e:
                erros += 1
                resultados.append({
                    'id': nota_id,
                    'sucesso': False,
                    'mensagem': f'Erro: {str(e)}'
                })
        
        # Mensagem de resumo profissional
        if emitidas > 0 and erros == 0:
            mensagem_resumo = f'🎉 Processamento concluído com sucesso!\n\nTodas as {emitidas} nota(s) foram emitidas corretamente.'
        elif emitidas > 0 and erros > 0:
            mensagem_resumo = f'⚠️ Processamento concluído com avisos.\n\n{emitidas} nota(s) emitida(s) com sucesso.\n{erros} nota(s) com erro.'
        else:
            mensagem_resumo = f'❌ Não foi possível emitir as notas.\n\n{erros} erro(s) encontrado(s).'
        
        return JsonResponse({
            'sucesso': emitidas > 0,
            'mensagem': mensagem_resumo,
            'emitidas': emitidas,
            'erros': erros,
            'total': len(notas_ids),
            'resultados': resultados
        })
    
    return redirect('nfs_sp:emitir')


@login_required
def excluir_notas(request):
    """Exclui notas pendentes selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            messages.error(request, 'Nenhuma nota selecionada.')
            return redirect('nfs_sp:emitir')
        
        # Excluir apenas pendentes
        excluidas = NotaFiscalSP.objects.filter(
            id__in=notas_ids,
            status_rps='pendente'
        ).delete()[0]
        
        if excluidas > 0:
            messages.success(request, f'{excluidas} nota(s) excluída(s) com sucesso.')
        else:
            messages.warning(request, 'Nenhuma nota pendente foi excluída.')
    
    return redirect('nfs_sp:emitir')


@login_required
def cancelar_notas(request):
    """Cancela notas (RPS) emitidas selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma nota selecionada.'
            })
        
        # Processar cancelamento
        resultados = []
        canceladas = 0
        erros = 0
        
        for nota_id in notas_ids:
            try:
                nota = NotaFiscalSP.objects.get(id=nota_id)
                
                # Verificar se está emitida
                if nota.status_rps != 'emitida':
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Não pode ser cancelado:\n\n'
                                   f'Status atual: {nota.get_status_rps_display()}\n'
                                   f'Apenas notas emitidas podem ser canceladas.\n\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    erros += 1
                    continue
                
                # Verificar se tem número da NFS-e
                if not nota.numero_nfse:
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Não pode ser cancelado:\n\n'
                                   f'Número da NFS-e não encontrado.\n'
                                   f'A nota pode não ter sido emitida corretamente.\n\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    erros += 1
                    continue
                
                # Criar XML de cancelamento e enviar
                evento = EventoNFeDjango(nota.empresa)
                xml = evento.criar_pedido_cancelamento_nfe(nota)
                
                processador = ProcessadorNFeDjango(nota.empresa)
                resultado = processador.cancelar_nfe(xml)
                
                if resultado['sucesso']:
                    # Atualizar nota
                    nota.status_rps = 'cancelada'
                    nota.save()
                    
                    canceladas += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': True,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'✅ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Cancelado com sucesso!\n'
                                   f'NFS-e: {nota.numero_nfse}\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                else:
                    erros += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_rps': nota.numero_rps,
                        'serie_rps': nota.serie_rps,
                        'tomador': nota.nome_tomador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ RPS Nº {nota.numero_rps} - Série {nota.serie_rps}\n'
                                   f'Erro no cancelamento:\n\n'
                                   f'{resultado.get("mensagem", "Erro ao cancelar RPS")}\n\n'
                                   f'Tomador: {nota.nome_tomador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    
            except Exception as e:
                erros += 1
                resultados.append({
                    'id': nota_id,
                    'sucesso': False,
                    'mensagem': f'Erro: {str(e)}'
                })
        
        # Mensagem de resumo profissional
        if canceladas > 0 and erros == 0:
            mensagem_resumo = f'🎉 Cancelamento concluído com sucesso!\n\nTodas as {canceladas} nota(s) foram canceladas corretamente.'
        elif canceladas > 0 and erros > 0:
            mensagem_resumo = f'⚠️ Cancelamento concluído com avisos.\n\n{canceladas} nota(s) cancelada(s) com sucesso.\n{erros} nota(s) com erro.'
        else:
            mensagem_resumo = f'❌ Não foi possível cancelar as notas.\n\n{erros} erro(s) encontrado(s).'
        
        return JsonResponse({
            'sucesso': canceladas > 0,
            'mensagem': mensagem_resumo,
            'canceladas': canceladas,
            'erros': erros,
            'total': len(notas_ids),
            'resultados': resultados
        })
    
    return redirect('nfs_sp:emitir')


@login_required
def salvar_pdfs(request):
    """Salva PDFs das notas emitidas"""
    notas_ids = request.GET.get('notas', '').split(',')
    
    if not notas_ids or notas_ids == ['']:
        messages.error(request, 'Nenhuma nota selecionada.')
        return redirect('nfs_sp:emitir')
    
    # TODO: Implementar download de PDFs
    messages.info(request, f'Download de PDFs em desenvolvimento.')
    
    return redirect('nfs_sp:emitir')


@login_required
def detalhes_nota(request, nota_id):
    """Retorna HTML com detalhes da nota"""
    nota = get_object_or_404(NotaFiscalSP, id=nota_id)
    
    html = f"""
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Tomador</h6>
            <p><strong>Nome:</strong> {nota.nome_tomador}</p>
            <p><strong>CNPJ/CPF:</strong> {nota.cnpj_cpf_tomador}</p>
            <p><strong>E-mail:</strong> {nota.email_tomador or '-'}</p>
            <p><strong>Endereço:</strong> {nota.logradouro_tomador or ''} {nota.numero_tomador or ''}</p>
            <p><strong>Bairro:</strong> {nota.bairro_tomador or '-'}</p>
            <p><strong>Cidade/UF:</strong> {nota.cidade_tomador or '-'}/{nota.uf_tomador or '-'}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Serviço</h6>
            <p><strong>Código:</strong> {nota.cod_servico}</p>
            <p><strong>Descrição:</strong> {nota.descricao}</p>
            <p><strong>Valor Total:</strong> R$ {nota.valor_total}</p>
            <p><strong>Deduções:</strong> R$ {nota.deducoes}</p>
            <p><strong>Alíquota:</strong> {nota.aliquota}%</p>
            <p><strong>Valor ISS:</strong> R$ {nota.valor_iss:.2f}</p>
        </div>
    </div>
    <hr>
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Retenções</h6>
            <p><strong>ISS Retido:</strong> {'Sim' if nota.iss_retido else 'Não'}</p>
            <p><strong>PIS:</strong> R$ {nota.pis_retido}</p>
            <p><strong>COFINS:</strong> R$ {nota.cofins_retido}</p>
            <p><strong>IRRF:</strong> R$ {nota.irrf_retido}</p>
            <p><strong>CSLL:</strong> R$ {nota.csll_retido}</p>
            <p><strong>INSS:</strong> R$ {nota.inss_retido}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Controle</h6>
            <p><strong>Status:</strong> {nota.get_status_rps_display()}</p>
            <p><strong>Data Emissão:</strong> {nota.data_emissao or '-'}</p>
            <p><strong>Número NFS-e:</strong> {nota.numero_nfse or '-'}</p>
            <p><strong>Código Verificação:</strong> {nota.codigo_verificacao or '-'}</p>
            <p><strong>Valor Líquido:</strong> R$ {nota.valor_liquido:.2f}</p>
        </div>
    </div>
    """
    
    return HttpResponse(html)


def consultar_notas_api(request):
    """Consulta notas na API da Prefeitura"""
    # TODO: Implementar integração com API
    messages.info(request, 'Consulta de notas via API em desenvolvimento.')
    return redirect('nfs_sp:emitir')


def gerar_relatorio(request):
    """Gera relatório conforme tipo selecionado"""
    # TODO: Implementar geração de relatórios
    messages.info(request, 'Geração de relatórios em desenvolvimento.')
    return redirect('nfs_sp:emitir')


# ========== VIEWS PARA NFTS (Nota Fiscal do Tomador) ==========

@login_required
def gerar_modelo_nfts(request):
    """Gera arquivo Excel modelo profissional para importação de NFTS"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo NFTS SP"
    
    # Cabeçalhos
    headers = [
        'CNPJ_TOMADOR', 'INSCRIÇÃO_MUNICIPAL', 'DATA_PRESTAÇÃO_SERVIÇO',
        'CNPJ_CPF_PRESTADOR', 'NUMERO_DOCUMENTO', 'SERIE', 'CIDADE', 'ESTADO', 'CEP',
        'COD_SERVIÇO', 'DESCRIÇÃO', 'VALOR_TOTAL', 'DEDUÇÕES', 'ALIQUOTA',
        'TIPO_TRIBUTAÇÃO', 'REGIME_TRIBUTAÇÃO', 'TIPO_DOCUMENTO', 'NFTS', 'STATUS_NFTS', 'ISS_RETIDO'
    ]
    
    # Estilizar cabeçalho com cores Tax Gold
    header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    header_font = Font(bold=True, color='000000', size=11, name='Calibri')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Aplicar estilo ao cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar altura da linha do cabeçalho
    ws.row_dimensions[1].height = 30
    
    # Linha de exemplo com dados realistas
    exemplo = [
        '12345678000190',  # CNPJ_TOMADOR
        '123456789',  # INSCRIÇÃO_MUNICIPAL
        '2024-11-20',  # DATA_PRESTAÇÃO_SERVIÇO
        '98765432000100',  # CNPJ_CPF_PRESTADOR
        '12345',  # NUMERO_DOCUMENTO
        '001',  # SERIE
        'São Paulo',  # CIDADE
        'SP',  # ESTADO
        '01310-100',  # CEP
        '07498',  # COD_SERVIÇO
        'Serviços de consultoria empresarial e gestão',  # DESCRIÇÃO
        '5000.00',  # VALOR_TOTAL
        '0.00',  # DEDUÇÕES
        '2.00',  # ALIQUOTA
        'T',  # TIPO_TRIBUTAÇÃO (T=Tributado, F=Fora, I=Isento, N=Não Tributado)
        'simples',  # REGIME_TRIBUTAÇÃO (simples, presumido, real, mei)
        'nfse',  # TIPO_DOCUMENTO (nfe, nfse, cupom, recibo)
        '',  # NFTS (vazio para importação)
        'Pendente',  # STATUS_NFTS
        'SIM'  # ISS_RETIDO
    ]
    
    # Adicionar linha de exemplo com estilo
    for col_num, valor in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col_num, value=valor)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(name='Calibri', size=10)
    
    # Adicionar linha de instruções (comentário)
    instrucoes_row = 4
    ws.merge_cells(f'A{instrucoes_row}:T{instrucoes_row}')
    cell_instrucoes = ws.cell(row=instrucoes_row, column=1)
    cell_instrucoes.value = (
        "INSTRUÇÕES: Preencha os campos marcados com * (obrigatórios). "
        "TIPO_TRIBUTAÇÃO: T=Tributado, F=Fora do Município, I=Isento, N=Não Tributado. "
        "REGIME_TRIBUTAÇÃO: simples, presumido, real, mei. "
        "TIPO_DOCUMENTO: nfe, nfse, cupom, recibo. "
        "ISS_RETIDO: SIM ou NÃO"
    )
    cell_instrucoes.font = Font(italic=True, color='666666', size=9)
    cell_instrucoes.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[instrucoes_row].height = 40
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 18,  # CNPJ_TOMADOR
        'B': 20,  # INSCRIÇÃO_MUNICIPAL
        'C': 22,  # DATA_PRESTAÇÃO_SERVIÇO
        'D': 18,  # CNPJ_CPF_PRESTADOR
        'E': 18,  # NUMERO_DOCUMENTO
        'F': 10,  # SERIE
        'G': 15,  # CIDADE
        'H': 10,  # ESTADO
        'I': 12,  # CEP
        'J': 15,  # COD_SERVIÇO
        'K': 40,  # DESCRIÇÃO
        'L': 15,  # VALOR_TOTAL
        'M': 12,  # DEDUÇÕES
        'N': 12,  # ALIQUOTA
        'O': 18,  # TIPO_TRIBUTAÇÃO
        'P': 20,  # REGIME_TRIBUTAÇÃO
        'Q': 18,  # TIPO_DOCUMENTO
        'R': 15,  # NFTS
        'S': 15,  # STATUS_NFTS
        'T': 12,  # ISS_RETIDO
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Adicionar bordas às células
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="TaxGold_Modelo_NFTS_SP.xlsx"'
    wb.save(response)
    
    return response


def importar_planilha_nfts(request, empresas):
    """Importa planilha Excel com NFTS"""
    try:
        empresa_id = request.POST.get('empresa_id')
        arquivo = request.FILES.get('arquivo_excel')
        
        if not empresa_id or not arquivo:
            messages.error(request, 'Selecione a empresa e o arquivo Excel.')
            return redirect('nfs_sp:emitir')
        
        empresa = get_object_or_404(Empresa, id=empresa_id)
        
        # Validar extensão
        if not arquivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Arquivo deve ser .xlsx ou .xls')
            return redirect('nfs_sp:emitir')
        
        # Ler planilha
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        importadas = 0
        erros = []
        
        # Pular cabeçalho (linha 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Validar se linha não está vazia
                if not any(row):
                    continue
                
                # Converter data
                data_prestacao = row[2]
                if isinstance(data_prestacao, str):
                    data_prestacao = datetime.strptime(data_prestacao, '%Y-%m-%d').date()
                elif isinstance(data_prestacao, datetime):
                    data_prestacao = data_prestacao.date()
                
                # Extrair dados (conforme colunas especificadas)
                nota = NotaFiscalTomadorSP(
                    empresa=empresa,
                    cnpj_contribuinte=empresa.cnpj,  # CNPJ da empresa tomadora
                    cnpj_tomador=str(row[0] or '').strip(),
                    inscricao_municipal=str(row[1] or '').strip(),
                    data_prestacao_servico=data_prestacao,
                    cnpj_cpf_prestador=str(row[3] or '').strip(),
                    numero_documento=str(row[4] or '').strip() if row[4] else None,
                    serie=str(row[5] or '').strip() if row[5] else None,
                    cidade=str(row[6] or '').strip() if row[6] else None,
                    estado=str(row[7] or '').strip() if row[7] else None,
                    cep=str(row[8] or '').strip() if row[8] else None,
                    cod_servico=str(row[9] or '').strip(),
                    descricao=str(row[10] or '').strip(),
                    valor_total=Decimal(str(row[11] or 0)),
                    deducoes=Decimal(str(row[12] or 0)),
                    aliquota=Decimal(str(row[13] or 0)),
                    tipo_tributacao=str(row[14] or 'T').strip()[0].upper(),
                    regime_tributacao=str(row[15] or 'simples').strip().lower(),
                    tipo_documento=str(row[16] or 'nfse').strip().lower(),
                    nfts=str(row[17] or '').strip() if row[17] else None,
                    iss_retido=str(row[19] or '').strip().upper() in ['SIM', 'S', 'TRUE', '1'],
                    status_nfts='pendente'
                )
                nota.save()
                importadas += 1
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        if importadas > 0:
            messages.success(request, f'{importadas} NFTS importada(s) com sucesso!')
        
        if erros:
            messages.warning(request, f'{len(erros)} erro(s) encontrado(s). Verifique os dados.')
        
    except Exception as e:
        messages.error(request, f'Erro ao importar planilha: {str(e)}')
    
    return redirect('nfs_sp:emitir')


@login_required
def emitir_nfts(request):
    """Emite NFTS selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma NFTS selecionada.'
            })
        
        # Processar emissão
        resultados = []
        emitidas = 0
        erros = 0
        
        for nota_id in notas_ids:
            try:
                nota = NotaFiscalTomadorSP.objects.get(id=nota_id)
                
                # Verificar se já foi emitida ou cancelada (não permite reenvio)
                if nota.status_nfts in ['emitida', 'cancelada']:
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_documento': nota.numero_documento,
                        'serie': nota.serie,
                        'prestador': nota.cnpj_cpf_prestador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ NFTS Nº {nota.numero_documento}' + (f' - Série {nota.serie}' if nota.serie else '') + f'\n'
                                   f'Não pode ser emitida:\n\n'
                                   f'Status atual: {nota.get_status_nfts_display()}\n'
                                   f'Esta nota já foi {nota.get_status_nfts_display().lower()}.\n\n'
                                   f'Prestador: {nota.cnpj_cpf_prestador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    erros += 1
                    continue
                
                # Se estava com erro, limpar mensagem de erro anterior
                if nota.status_nfts == 'erro':
                    nota.mensagem_erro = None
                
                # Criar XML e enviar
                evento = EventoNFeDjango(nota.empresa)
                xml = evento.criar_pedido_envio_nfts(nota)
                
                processador = ProcessadorNFeDjango(nota.empresa)
                resultado = processador.enviar_nfts(xml)
                
                if resultado['sucesso']:
                    # Atualizar nota
                    nota.nfts = resultado.get('numero_nfe')
                    nota.codigo_verificacao = resultado.get('codigo_verificacao')
                    nota.status_nfts = 'emitida'
                    nota.data_emissao = datetime.now()
                    nota.save()
                    
                    emitidas += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': True,
                        'numero_documento': nota.numero_documento,
                        'serie': nota.serie,
                        'prestador': nota.cnpj_cpf_prestador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'✅ NFTS Nº {nota.numero_documento}' + (f' - Série {nota.serie}' if nota.serie else '') + f'\n'
                                   f'Emitida com sucesso!\n'
                                   f'Prestador: {nota.cnpj_cpf_prestador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}',
                        'numero_nfe': nota.nfts,
                        'codigo_verificacao': nota.codigo_verificacao,
                        'url': resultado.get('url_nfe')
                    })
                else:
                    nota.status_nfts = 'erro'
                    nota.mensagem_erro = resultado.get('mensagem', 'Erro desconhecido')
                    nota.save()
                    
                    erros += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'numero_documento': nota.numero_documento,
                        'serie': nota.serie,
                        'prestador': nota.cnpj_cpf_prestador,
                        'valor': float(nota.valor_total),
                        'mensagem': f'❌ NFTS Nº {nota.numero_documento}' + (f' - Série {nota.serie}' if nota.serie else '') + f'\n'
                                   f'Erro na emissão:\n\n'
                                   f'{resultado.get("mensagem", "Erro ao emitir NFTS")}\n\n'
                                   f'Prestador: {nota.cnpj_cpf_prestador}\n'
                                   f'Valor: R$ {nota.valor_total:,.2f}'
                    })
                    
            except Exception as e:
                erros += 1
                resultados.append({
                    'id': nota_id,
                    'sucesso': False,
                    'mensagem': f'Erro: {str(e)}'
                })
        
        # Mensagem de resumo profissional
        if emitidas > 0 and erros == 0:
            mensagem_resumo = f'🎉 Emissão concluída com sucesso!\n\nTodas as {emitidas} NFTS foram emitidas corretamente.'
        elif emitidas > 0 and erros > 0:
            mensagem_resumo = f'⚠️ Emissão concluída com avisos.\n\n{emitidas} NFTS emitida(s) com sucesso.\n{erros} NFTS com erro.'
        else:
            mensagem_resumo = f'❌ Não foi possível emitir as NFTS.\n\n{erros} erro(s) encontrado(s).'
        
        return JsonResponse({
            'sucesso': emitidas > 0,
            'mensagem': mensagem_resumo,
            'emitidas': emitidas,
            'erros': erros,
            'total': len(notas_ids),
            'resultados': resultados
        })
    
    return JsonResponse({'sucesso': False, 'mensagem': 'Método não permitido'})


@login_required
def excluir_nfts(request):
    """Exclui NFTS pendentes selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            messages.error(request, 'Nenhuma NFTS selecionada.')
            return redirect('nfs_sp:emitir')
        
        # Excluir apenas pendentes
        excluidas = NotaFiscalTomadorSP.objects.filter(
            id__in=notas_ids,
            status_nfts='pendente'
        ).delete()[0]
        
        if excluidas > 0:
            messages.success(request, f'{excluidas} NFTS excluída(s) com sucesso.')
        else:
            messages.warning(request, 'Nenhuma NFTS pendente foi excluída.')
    
    return redirect('nfs_sp:emitir')


@login_required
def cancelar_nfts(request):
    """Cancela NFTS emitidas selecionadas"""
    if request.method == 'POST':
        notas_ids = request.POST.getlist('notas[]')
        
        if not notas_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhuma NFTS selecionada.'
            })
        
        # Processar cancelamento
        resultados = []
        canceladas = 0
        erros = 0
        
        for nota_id in notas_ids:
            try:
                nota = NotaFiscalTomadorSP.objects.get(id=nota_id)
                
                # Verificar se pode ser cancelada
                if nota.status_nfts != 'emitida':
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'mensagem': f'NFTS {nota.numero_documento} não está emitida'
                    })
                    erros += 1
                    continue
                
                # Criar XML de cancelamento e enviar
                evento = EventoNFeDjango(nota.empresa)
                xml = evento.cancelamento_nfe(nota)
                
                processador = ProcessadorNFeDjango(nota.empresa)
                resultado = processador.cancelar_nfe(xml)
                
                if resultado['sucesso']:
                    # Atualizar nota
                    nota.status_nfts = 'cancelada'
                    nota.data_cancelamento = datetime.now()
                    nota.save()
                    
                    canceladas += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': True,
                        'mensagem': f'NFTS {nota.numero_documento} cancelada com sucesso!'
                    })
                else:
                    erros += 1
                    resultados.append({
                        'id': nota_id,
                        'sucesso': False,
                        'mensagem': resultado.get('mensagem', 'Erro ao cancelar NFTS')
                    })
                    
            except Exception as e:
                erros += 1
                resultados.append({
                    'id': nota_id,
                    'sucesso': False,
                    'mensagem': f'Erro: {str(e)}'
                })
        
        return JsonResponse({
            'sucesso': canceladas > 0,
            'mensagem': f'{canceladas} NFTS cancelada(s) com sucesso. {erros} erro(s).',
            'canceladas': canceladas,
            'erros': erros,
            'resultados': resultados
        })
    
    return JsonResponse({'sucesso': False, 'mensagem': 'Método não permitido'})


@login_required
def detalhes_nfts(request, nota_id):
    """Retorna HTML com detalhes da NFTS"""
    nota = get_object_or_404(NotaFiscalTomadorSP, id=nota_id)
    
    html = f"""
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Tomador (Declarante)</h6>
            <p><strong>CNPJ:</strong> {nota.cnpj_tomador}</p>
            <p><strong>Inscrição Municipal:</strong> {nota.inscricao_municipal}</p>
            <p><strong>Data Prestação:</strong> {nota.data_prestacao_servico.strftime('%d/%m/%Y')}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Prestador</h6>
            <p><strong>CNPJ/CPF:</strong> {nota.cnpj_cpf_prestador}</p>
            <p><strong>Documento:</strong> {nota.numero_documento or '-'} / Série: {nota.serie or '-'}</p>
            <p><strong>Cidade/UF:</strong> {nota.cidade or '-'}/{nota.estado or '-'}</p>
            <p><strong>CEP:</strong> {nota.cep or '-'}</p>
        </div>
    </div>
    <hr>
    <div class="row">
        <div class="col-md-6">
            <h6 class="text-primary">Dados do Serviço</h6>
            <p><strong>Código:</strong> {nota.cod_servico}</p>
            <p><strong>Descrição:</strong> {nota.descricao}</p>
            <p><strong>Valor Total:</strong> R$ {nota.valor_total}</p>
            <p><strong>Deduções:</strong> R$ {nota.deducoes}</p>
            <p><strong>Alíquota:</strong> {nota.aliquota}%</p>
            <p><strong>Valor ISS:</strong> R$ {nota.valor_iss:.2f}</p>
        </div>
        <div class="col-md-6">
            <h6 class="text-primary">Controle e Tributação</h6>
            <p><strong>Tipo Tributação:</strong> {nota.get_tipo_tributacao_display()}</p>
            <p><strong>Regime:</strong> {nota.get_regime_tributacao_display()}</p>
            <p><strong>Tipo Documento:</strong> {nota.get_tipo_documento_display()}</p>
            <p><strong>ISS Retido:</strong> {'Sim' if nota.iss_retido else 'Não'}</p>
            <p><strong>Status:</strong> {nota.get_status_nfts_display()}</p>
            <p><strong>NFTS:</strong> {nota.nfts or '-'}</p>
        </div>
    </div>
    """
    
    return HttpResponse(html)
